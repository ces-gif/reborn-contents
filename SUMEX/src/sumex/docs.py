"""서류 생성 — 거래명세서 / 가납서 / 선납서 / 견적서.

두 가지 모드로 동작한다.

  1) templates/ 에 회사 실양식 xlsx 가 있으면 그 파일을 열어 값만 채운다.
     (로고·서식·인쇄 설정이 그대로 유지되므로 이 쪽이 원칙)
  2) 없으면 doc_types.yaml 의 좌표대로 동일한 레이아웃을 새로 만든다.
     양식 파일 없이도 바로 쓸 수 있게 하기 위한 대비책이다.

금액 규칙은 실제 자사 양식과 같다.
  단가·금액은 VAT 포함가, 공급가액 = 합계/1.1, 부가세 = 합계 - 공급가액.
"""
from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from . import config, registry
from .registry import Hospital

ITEM_FIRST_ROW = 12
ITEM_LAST_ROW = 37
TOTAL_ROW = 38
FONT = "맑은 고딕"

_SAFE = re.compile(r'[\\/:*?"<>|]+')


@dataclass
class Item:
    name: str
    qty: float
    price: float = 0.0
    code: str = ""
    unit: str = "EA"

    @property
    def amount(self) -> float:
        return round(self.qty * self.price)


def parse_items(source: str | Path | Sequence[dict[str, Any]]) -> list[Item]:
    """CSV 경로 / JSON 경로 / dict 리스트 / 'name x qty @ price' 문자열을 Item 으로."""
    if isinstance(source, (list, tuple)):
        return [_item_from_dict(dict(row)) for row in source]

    text = str(source)
    p = _as_existing_file(text)
    if p is not None:
        if p.suffix.lower() == ".json":
            return [_item_from_dict(r) for r in json.loads(p.read_text(encoding="utf-8"))]
        with p.open(encoding="utf-8-sig", newline="") as fh:
            return [_item_from_dict(r) for r in csv.DictReader(fh)]

    # "ICONIX 1.7T x 3 @ 320000; ICONIX NEEDLES x 2 @ 180000"
    out: list[Item] = []
    for chunk in re.split(r"[;\n]", text):
        chunk = chunk.strip()
        if not chunk:
            continue
        m = re.match(r"^(?P<name>.+?)\s*[xX*]\s*(?P<qty>[\d.]+)\s*(?:@\s*(?P<price>[\d,]+))?$", chunk)
        if not m:
            raise ValueError(
                f"품목을 해석하지 못했습니다: {chunk!r}\n"
                "형식: '품목명 x 수량 @ 단가'  (예: 'ICONIX 1.7T x 3 @ 320000')"
            )
        out.append(Item(
            name=m.group("name").strip(),
            qty=float(m.group("qty")),
            price=float((m.group("price") or "0").replace(",", "")),
        ))
    if not out:
        raise ValueError("품목이 비어 있습니다.")
    return out


def _as_existing_file(text: str) -> Path | None:
    """문자열이 '실제로 존재하는 파일 경로'일 때만 Path 를 준다.

    빈 문자열이나 아주 긴 인라인 품목 목록을 경로로 오해하면 OSError 가 난다.
    """
    candidate = text.strip()
    if not candidate or len(candidate) > 255 or "\n" in candidate or ";" in candidate:
        return None
    try:
        p = Path(candidate)
        return p if p.is_file() else None
    except OSError:
        return None


def _item_from_dict(row: dict[str, Any]) -> Item:
    def pick(*keys: str, default: Any = "") -> Any:
        for k in keys:
            if k in row and row[k] not in (None, ""):
                return row[k]
        return default

    return Item(
        name=str(pick("name", "상품명", "품목", "제품명")).strip(),
        qty=float(str(pick("qty", "수량", default=1)).replace(",", "")),
        price=float(str(pick("price", "단가", default=0)).replace(",", "")),
        code=str(pick("code", "코드", "모델명")).strip(),
        unit=str(pick("unit", "포장단위", default="EA")).strip(),
    )


def filename(doc_type: str, subject: str, when: date, suffix: str = "") -> str:
    pattern = registry.filename_pattern()
    name = pattern.format(
        doc_type=doc_type,
        subject=subject,
        yymmdd=when.strftime("%y%m%d"),
        suffix=f"_{suffix}" if suffix else "",
    )
    return _SAFE.sub("_", name)


# ── 레이아웃 ────────────────────────────────────────────────
_THIN = Side(style="thin", color="000000")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _blank_statement(title: str) -> Workbook:
    """실양식이 없을 때 쓰는 동일 좌표 레이아웃."""
    wb = Workbook()
    ws = wb.active
    ws.title = "견적서 (2)"

    widths = {"A": 5, "B": 13, "C": 34, "D": 10, "E": 7, "F": 12, "G": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells("A2:G2")
    cell = ws["A2"]
    cell.value = title
    cell.font = Font(name=FONT, size=20, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 34

    for coord in ("A7:C7", "A8:C8", "A9:C9"):
        ws.merge_cells(coord)

    headers = ["NO", "코 드", "상 품 명", "포장단위", "수량", "단 가", "금 액"]
    for idx, text in enumerate(headers, start=1):
        c = ws.cell(row=11, column=idx, value=text)
        c.font = Font(name=FONT, size=10, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BOX

    ws.merge_cells(f"A{TOTAL_ROW}:E{TOTAL_ROW}")
    ws.merge_cells("A39:E41")
    return wb


def _load_template(doc_type: str) -> tuple[Workbook, bool]:
    for candidate in (f"{doc_type}.xlsx", "거래명세서.xlsx"):
        p = config.path("templates", candidate)
        if p.exists():
            return load_workbook(p), True
    return _blank_statement(" ".join(doc_type)), False


def build_statement(
    hospital: str | Hospital,
    items: Iterable[Item],
    *,
    doc_type: str = "거래명세서",
    when: date | None = None,
    subject: str | None = None,
    suffix: str = "",
    out_path: Path | None = None,
) -> Path:
    """거래명세서·가납서·선납서를 만든다. 반환값은 저장된 경로."""
    h = hospital if isinstance(hospital, Hospital) else registry.find(hospital)
    when = when or date.today()
    items = list(items)
    if not items:
        raise ValueError("품목이 하나도 없습니다.")
    if len(items) > ITEM_LAST_ROW - ITEM_FIRST_ROW + 1:
        raise ValueError(
            f"한 장에 들어갈 수 있는 품목은 최대 {ITEM_LAST_ROW - ITEM_FIRST_ROW + 1}건입니다. "
            f"({len(items)}건 요청) — 여러 장으로 나눠 발행하세요."
        )

    cfg = config.load()
    company = cfg.get("company", {})

    wb, from_template = _load_template(doc_type)
    ws = wb["견적서 (2)"] if "견적서 (2)" in wb.sheetnames else wb.worksheets[0]

    ws["D6"] = f"사업자등록번호 : {company.get('biz_no')}"
    ws["A7"] = when
    ws["A7"].number_format = "yyyy-mm-dd"
    ws["D7"] = f"상호명  : {company.get('name')}"
    ws["D8"] = f"주 소 : {company.get('address')}"
    ws["A9"] = f"병원명 : {h.name}"
    ws["D9"] = f"대표자 : {company.get('ceo')}"

    for offset, item in enumerate(items):
        row = ITEM_FIRST_ROW + offset
        ws.cell(row=row, column=1, value=offset + 1)
        ws.cell(row=row, column=2, value=item.code)
        ws.cell(row=row, column=3, value=item.name)
        ws.cell(row=row, column=4, value=item.unit)
        ws.cell(row=row, column=5, value=item.qty)
        ws.cell(row=row, column=6, value=item.price)
        ws.cell(row=row, column=7, value=f"=E{row}*F{row}")
        for col in range(1, 8):
            c = ws.cell(row=row, column=col)
            c.border = _BOX
            c.font = Font(name=FONT, size=10)
            if col in (5, 6, 7):
                c.number_format = "#,##0"

    # 사용하지 않는 품목 행은 비워둔다 (템플릿 재사용 시 잔여값 제거)
    for row in range(ITEM_FIRST_ROW + len(items), ITEM_LAST_ROW + 1):
        for col in range(1, 8):
            ws.cell(row=row, column=col, value=None)

    ws[f"F{TOTAL_ROW}"] = f"=SUM(G{ITEM_FIRST_ROW}:G{ITEM_LAST_ROW})"
    ws[f"G{TOTAL_ROW}"] = "(VAT 포함)"
    ws["F39"], ws["G39"] = "공급가액", f"=F{TOTAL_ROW}/1.1"
    ws["F40"], ws["G40"] = "부가세", "=G41-G39"
    ws["F41"], ws["G41"] = "합 계", f"=F{TOTAL_ROW}"
    for coord in ("G39", "G40", "G41", f"F{TOTAL_ROW}"):
        ws[coord].number_format = "#,##0"
    ws["C42"] = f"결제계좌 : {company.get('bank_account')}"

    if not from_template:
        for coord in ("A7", "D6", "D7", "D8", "A9", "D9", "C42", "F39", "F40", "F41"):
            ws[coord].font = Font(name=FONT, size=10)

    subject = subject or _default_subject(items)
    target = out_path or config.out_dir("서류") / filename(doc_type, subject, when, suffix)
    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    return target


def _default_subject(items: Sequence[Item]) -> str:
    if len(items) == 1:
        return items[0].name
    head = items[0].name.split()[0] if items[0].name else "품목"
    return f"{head} 외 {len(items) - 1}건"


def build_quotation(
    hospital: str | Hospital,
    groups: Sequence[tuple[str, Sequence[Item]]],
    *,
    when: date | None = None,
    subject: str | None = None,
    out_path: Path | None = None,
) -> Path:
    """견적서. groups 는 [("1788 VIDEO CAMERA SYSTEM", [Item, ...]), ...]."""
    h = hospital if isinstance(hospital, Hospital) else registry.find(hospital)
    when = when or date.today()
    cfg = config.load()
    company, rep = cfg.get("company", {}), cfg.get("rep", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "견적서"
    for col, width in {"A": 5, "B": 18, "C": 40, "D": 7, "E": 16, "F": 16}.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells("A2:F2")
    ws["A2"] = "Q U O T A T I O N"
    ws["A2"].font = Font(name=FONT, size=20, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["E3"] = f"등록번호 : {company.get('biz_no')}"
    ws["E4"] = f"상     호 : {company.get('name')}"
    ws["E5"] = f"대     표 : {company.get('ceo')}"
    ws["E6"] = f"주     소 : {company.get('address')}"

    ws.merge_cells("A5:C5")
    ws["A5"] = when
    ws["A5"].number_format = "yyyy-mm-dd"
    ws.merge_cells("A7:C7")
    ws["A7"] = f"{h.name} 귀하"

    for idx, text in enumerate(["", "  CAT. NO.", "DESCRIPTION", "QTY", "UNIT.PRICE(+VAT)", "AMOUNT"], start=1):
        c = ws.cell(row=10, column=idx, value=text)
        c.font = Font(name=FONT, size=10, bold=True)
        c.alignment = Alignment(horizontal="center")
        c.border = _BOX

    row = 11
    seq = 0
    for group_name, group_items in groups:
        if group_name:
            ws.cell(row=row, column=3, value=f"<{group_name}>").font = Font(name=FONT, size=10, bold=True)
            row += 1
        for item in group_items:
            seq += 1
            ws.cell(row=row, column=1, value=seq)
            ws.cell(row=row, column=2, value=item.code)
            ws.cell(row=row, column=3, value=item.name)
            ws.cell(row=row, column=4, value=item.qty)
            ws.cell(row=row, column=5, value=item.price)
            ws.cell(row=row, column=6, value=f"=D{row}*E{row}")
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = _BOX
                cell.font = Font(name=FONT, size=10)
                if col in (4, 5, 6):
                    cell.number_format = "#,##0"
            row += 1

    total_row = max(row + 1, 43)
    ws.merge_cells(f"A{total_row}:C{total_row}")
    ws[f"A{total_row}"] = "Total"
    ws[f"D{total_row}"] = f"=SUM(D11:D{row - 1})"
    ws[f"F{total_row}"] = f"=SUM(F11:F{row - 1})"
    ws[f"F{total_row}"].number_format = "#,##0"
    ws.merge_cells("A9:B9")
    ws["A9"] = "합  계"
    ws["C9"] = f"=F{total_row}"
    ws["C9"].number_format = "#,##0"

    ws[f"A{total_row + 1}"] = "◆"
    ws[f"B{total_row + 1}"] = "상기 견적가는 제출일로 부터 1개월간 유효합니다."
    ws[f"A{total_row + 2}"] = "◆"
    ws[f"E{total_row + 2}"] = f"영업담당자 : {rep.get('name')} {rep.get('title')} {rep.get('phone')}"

    subject = subject or (groups[0][0] if groups and groups[0][0] else "견적")
    target = out_path or config.out_dir("서류") / filename("견적서", f"{subject} ({h.short})", when)
    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    return target


def summarize(items: Sequence[Item]) -> dict[str, float]:
    total = sum(i.amount for i in items)
    supply = round(total / 1.1)
    return {"total": total, "supply": supply, "vat": total - supply}
