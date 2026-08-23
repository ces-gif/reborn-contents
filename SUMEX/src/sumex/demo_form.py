"""데모인수증 (KQF-OPC-009-F Korea Demo Receipt Form).

한국스트라이커 양식이므로 원칙은 회사가 배포한 xlsx 를 채우는 것이다.
templates/데모인수증.xlsx 가 있으면 그 파일의 값만 바꾸고, 없으면 같은 항목을
가진 작업용 시트를 만들어 준다. 어느 쪽이든 '데모 전 점검표'를 함께 낸다.

회수일은 기본값이 출고일 + 30일이다 (실제 사용 이력 기준).
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path
from typing import Any, Sequence

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

from . import config, registry
from .registry import Hospital

FONT = "맑은 고딕"
_THIN = Side(style="thin", color="000000")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
DEFAULT_LOAN_DAYS = 30


@dataclass
class DemoItem:
    model: str                 # 모델명 (예: 1788010000i)
    name: str                  # 제품명
    qty: int = 1
    serial: str = ""           # 제조번호 (Serial / Lot)
    license_no: str = "N/A"    # 허가번호 (예: 수신23-2838)
    category: str = "N/A"      # 품목명 (예: 의료영상 처리장치)
    maker: str = "N/A"         # 제조원상호
    is_device: str = "O"       # 의료기기 여부 O / X

    @property
    def registered(self) -> bool:
        return self.license_no not in ("", "N/A", "n/a")


@dataclass
class DemoRequest:
    hospital: Hospital
    dept: str
    doctor: str = ""              # 개인정보 — 서류에만 들어가고 저장소에는 남기지 않는다
    release_date: date = field(default_factory=date.today)
    return_date: date | None = None
    institution_no: str = ""      # 요양기관번호
    address: str = ""
    items: Sequence[DemoItem] = ()
    ship_method: str = "용달"
    ship_request_date: date | None = None

    def __post_init__(self) -> None:
        if self.return_date is None:
            self.return_date = self.release_date + timedelta(days=DEFAULT_LOAN_DAYS)
        if self.ship_request_date is None:
            self.ship_request_date = self.release_date - timedelta(days=1)


def _doc_spec() -> dict[str, Any]:
    return registry.doc_type("demo_receipt") or {}


def checklist(req: DemoRequest) -> list[str]:
    """데모 전에 반드시 확인할 것 + 이 요청에서 실제로 비어 있는 것."""
    spec = _doc_spec()
    lines = list(spec.get("checklist_before_demo") or [])

    lines.append(f"회수일 {req.return_date:%Y-%m-%d} — 캘린더에 회수 일정 등록")
    lines.append(f"회수 주소: {spec.get('return_address', '')}")
    lines.append("출고일 포함 3일 안에 제품·수량 확인. 3일 지나면 인수 완료로 처리된다")

    missing: list[str] = []
    if not req.institution_no:
        missing.append("요양기관번호")
    if not req.doctor:
        missing.append("의사명")
    if not req.address:
        missing.append("병원 주소")
    for item in req.items:
        if not item.serial:
            missing.append(f"제조번호(Serial/Lot) — {item.name}")
    if missing:
        lines.append("★ 빈 칸: " + ", ".join(dict.fromkeys(missing)))

    unregistered = [i.name for i in req.items if not i.registered and i.is_device == "O"]
    if unregistered:
        lines.append("? 의료기기인데 허가번호가 비어 있음: " + ", ".join(unregistered))

    return lines


def build(req: DemoRequest, out_path: Path | None = None) -> Path:
    cfg = config.load()
    rep = cfg.get("rep", {})
    spec = _doc_spec()

    template = config.path("templates", "데모인수증.xlsx")
    if template.exists():
        wb = load_workbook(template)
        ws = wb.worksheets[0]
        _fill_template(ws, req, rep)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "데모인수증"
        _build_fresh(ws, req, rep, spec)

    subject = f"{req.hospital.short}"
    if req.items:
        subject += f" ({req.items[0].name.split()[0]})"
    target = out_path or config.out_dir("데모") / (
        f"KQF-OPC-009-F Korea Demo Receipt Form_{subject}.xlsx".replace("/", "_")
    )
    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    return target


def _fill_template(ws: Any, req: DemoRequest, rep: dict[str, Any]) -> None:
    """실양식은 셀 좌표가 회사 배포본에 따라 달라질 수 있어, 라벨을 찾아 오른쪽에 채운다."""
    wanted = {
        "병원명": req.hospital.name,
        "요양기관번호": req.institution_no,
        "병원과": req.dept,
        "의사명": req.doctor,
        "출고일": req.release_date,
        "회수일": req.return_date,
        "담당자": f"{rep.get('name')} {rep.get('phone')}",
        "부서 및 대리점": "SUMEX",
    }
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40)):
        for cell in row:
            label = str(cell.value).strip() if cell.value else ""
            if label in wanted:
                _write_right_of(ws, cell, wanted[label])


def _write_right_of(ws: Any, cell: Any, value: Any) -> None:
    """라벨 셀의 오른쪽 첫 빈 칸(또는 기존 값 자리)에 쓴다."""
    for offset in range(1, 8):
        target = ws.cell(row=cell.row, column=cell.column + offset)
        if isinstance(target, type(cell)) and not _is_merged_child(ws, target):
            target.value = value
            if isinstance(value, date):
                target.number_format = "yyyy-mm-dd"
            return


def _is_merged_child(ws: Any, cell: Any) -> bool:
    for rng in ws.merged_cells.ranges:
        if cell.coordinate in rng and cell.coordinate != rng.coord.split(":")[0]:
            return True
    return False


def _build_fresh(ws: Any, req: DemoRequest, rep: dict[str, Any], spec: dict[str, Any]) -> None:
    for col, width in {"A": 6, "B": 20, "C": 42, "D": 8, "E": 20, "F": 16, "G": 24, "H": 22, "I": 12}.items():
        ws.column_dimensions[col].width = width

    def head(row: int, text: str) -> None:
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name=FONT, size=12, bold=True)

    def pair(row: int, label: str, value: Any) -> None:
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10, bold=True)
        c = ws.cell(row=row, column=2, value=value)
        c.font = Font(name=FONT, size=10)
        if isinstance(value, date):
            c.number_format = "yyyy-mm-dd"

    ws.merge_cells("A1:I1")
    ws["A1"] = f"데모인수증  Demo Receipt   ({spec.get('form_no', 'KQF-OPC-009-F')})"
    ws["A1"].font = Font(name=FONT, size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["E2"] = f"수입자 : {spec.get('issuer', '한국스트라이커㈜')}"

    head(4, "Demo Release Request (Stryker)")
    pair(5, "부서 및 대리점", "SUMEX")
    pair(6, "담당자", f"{rep.get('name')} {rep.get('phone')}")

    head(8, "Demo Receiver")
    pair(9, "병원명", req.hospital.name)
    pair(10, "요양기관번호", req.institution_no)
    pair(11, "주소", req.address)
    pair(12, "병원과", req.dept)
    pair(13, "의사명", req.doctor)
    pair(14, "출고일", req.release_date)
    pair(15, "회수일", req.return_date)

    head(17, "Product List")
    headers = ["No.", "모델명", "제품명", "Q'ty", "제조번호(Serial/Lot)", "허가번호", "품목명", "제조원상호", "의료기기여부"]
    for idx, text in enumerate(headers, start=1):
        c = ws.cell(row=18, column=idx, value=text)
        c.font = Font(name=FONT, size=9, bold=True)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = _BOX

    row = 19
    for idx, item in enumerate(req.items, start=1):
        values = [idx, item.model, item.name, item.qty, item.serial,
                  item.license_no, item.category, item.maker, item.is_device]
        for col, value in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(name=FONT, size=9)
            c.border = _BOX
        row += 1

    ws.cell(row=row, column=1, value="Total").font = Font(name=FONT, size=9, bold=True)
    ws.cell(row=row, column=4, value=sum(i.qty for i in req.items)).font = Font(name=FONT, size=9, bold=True)
    row += 2

    for idx, term in enumerate(spec.get("terms") or [], start=1):
        ws.cell(row=row, column=1, value=f"{idx}. {term}").font = Font(name=FONT, size=9)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value=f"회수 주소 : {spec.get('return_address', '')}").font = Font(name=FONT, size=9)
    row += 2

    for label in ("Date of Receive (의료기관 제공일자/설치완료일자)", "Date of Return (의료기관 회수일자)"):
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=9, bold=True)
        ws.cell(row=row, column=4, value="년    월    일")
        ws.cell(row=row, column=6, value="의료기관")
        ws.cell(row=row, column=7, value=req.hospital.name)
        ws.cell(row=row, column=9, value="(signature)")
        row += 1
        ws.cell(row=row, column=6, value="직판 / 대리점")
        ws.cell(row=row, column=7, value="SUMEX")
        ws.cell(row=row, column=9, value="(signature)")
        row += 2

    ws.cell(row=row, column=1, value="배송 관련").font = Font(name=FONT, size=10, bold=True)
    row += 1
    for label, value in (
        ("접수 담당자", f"{rep.get('name')} {rep.get('phone')}"),
        ("출고 요청일", req.ship_request_date),
        ("출고방법", req.ship_method),
        ("도착지 상세 주소", req.address or req.hospital.name),
        ("받는이", f"{rep.get('name')} {rep.get('phone')}"),
    ):
        pair(row, label, value)
        row += 1


def load_request(path: str | Path) -> DemoRequest:
    """yaml 로 데모 요청을 정의해 두고 불러 쓴다."""
    data = yaml.safe_load(Path(path).read_text(encoding="utf-8")) or {}
    items = [DemoItem(**row) for row in (data.get("items") or [])]
    return DemoRequest(
        hospital=registry.find(str(data["hospital"])),
        dept=str(data.get("dept", "")),
        doctor=str(data.get("doctor", "")),
        release_date=data.get("release_date") or date.today(),
        return_date=data.get("return_date"),
        institution_no=str(data.get("institution_no", "")),
        address=str(data.get("address", "")),
        items=items,
        ship_method=str(data.get("ship_method", "용달")),
    )
