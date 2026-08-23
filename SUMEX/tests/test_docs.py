"""서류 생성 — 금액 계산과 xlsx 레이아웃."""
from datetime import date

import openpyxl
import pytest

from sumex import docs


def test_parse_inline_items():
    items = docs.parse_items("ICONIX 1.7T x 3 @ 320000; ICONIX 1 NEEDLES x 2 @ 33350")
    assert len(items) == 2
    assert items[0].name == "ICONIX 1.7T"
    assert items[0].qty == 3
    assert items[0].amount == 960000


def test_parse_rejects_garbage():
    with pytest.raises(ValueError, match="해석하지 못했습니다"):
        docs.parse_items("이건 형식이 아니다")


def test_parse_rejects_empty():
    with pytest.raises(ValueError):
        docs.parse_items("")


def test_parse_csv(tmp_path):
    csv = tmp_path / "items.csv"
    csv.write_text("코드,상품명,포장단위,수량,단가\nB3340006,Exofin HVTA bond,EA,30,63620\n",
                   encoding="utf-8")
    items = docs.parse_items(csv)
    assert items[0].code == "B3340006"
    assert items[0].qty == 30


def test_vat_split_matches_real_invoice():
    """2026-08-14 세종스포츠 실제 명세표 합계 1,026,700원과 맞는지."""
    items = docs.parse_items("ICONIX 1.7T x 3 @ 320000; ICONIX 1 NEEDLES x 2 @ 33350")
    total = docs.summarize(items)
    assert total["total"] == 1026700
    assert total["supply"] + total["vat"] == total["total"]


def test_build_statement(tmp_path):
    items = docs.parse_items("ICONIX 1.7T x 3 @ 320000")
    out = tmp_path / "test.xlsx"
    docs.build_statement("세종스포츠", items, when=date(2026, 8, 14), out_path=out)

    ws = openpyxl.load_workbook(out).active
    assert ws["A2"].value == "거 래 명 세 서"
    assert "세종스포츠정형외과" in ws["A9"].value
    assert ws["C12"].value == "ICONIX 1.7T"
    assert ws["E12"].value == 3
    assert ws["G12"].value == "=E12*F12"
    assert ws["F38"].value == "=SUM(G12:G37)"
    assert ws["G39"].value == "=F38/1.1"


def test_statement_rejects_too_many_items(tmp_path):
    items = docs.parse_items("; ".join(f"품목{i} x 1 @ 1000" for i in range(30)))
    with pytest.raises(ValueError, match="최대"):
        docs.build_statement("세종스포츠", items, out_path=tmp_path / "x.xlsx")


def test_statement_rejects_empty(tmp_path):
    with pytest.raises(ValueError, match="품목이 하나도"):
        docs.build_statement("세종스포츠", [], out_path=tmp_path / "x.xlsx")


def test_filename_pattern():
    name = docs.filename("거래명세서", "엑소핀", date(2026, 8, 10), "수술방")
    assert name == "SUMEX 거래명세서(엑소핀)_260810_수술방.xlsx"


def test_filename_strips_path_separators():
    name = docs.filename("거래명세서", "A/B", date(2026, 1, 1))
    assert "/" not in name


def test_build_quotation(tmp_path):
    groups = [("1788 VIDEO CAMERA SYSTEM",
               docs.parse_items("1788 Camera Console x 1 @ 1000000"))]
    out = tmp_path / "q.xlsx"
    docs.build_quotation("호수병원", groups, when=date(2026, 9, 1), out_path=out)
    ws = openpyxl.load_workbook(out).active
    assert ws["A2"].value == "Q U O T A T I O N"
    assert "호수병원" in ws["A7"].value
    assert ws["C11"].value == "<1788 VIDEO CAMERA SYSTEM>"
